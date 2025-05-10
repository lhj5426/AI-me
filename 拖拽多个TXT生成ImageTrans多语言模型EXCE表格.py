import sys
import os
import openpyxl
from openpyxl.utils import get_column_letter

def read_txt_lines(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        # 去掉每行末尾的换行和 \t
        return [line.rstrip('\n\r').rstrip('\t') for line in f]

def main():
    txt_files = sys.argv[1:]
    if not txt_files:
        print("请将多个TXT文件拖拽到脚本上运行。")
        return

    wb = openpyxl.Workbook()
    ws = wb.active

    for col_index, txt_path in enumerate(txt_files, start=1):
        lines = read_txt_lines(txt_path)
        col_letter = get_column_letter(col_index)

        # 把文件名前8个字符写入表头
        filename = os.path.basename(txt_path)
        header = filename[:8].ljust(8)
        ws[f"{col_letter}1"] = header

        # 从第2行开始写入内容
        for row_index, line in enumerate(lines, start=2):
            ws[f"{col_letter}{row_index}"] = line

    # 自动生成保存路径
    base_dir = os.path.dirname(txt_files[0])
    output_path = os.path.join(base_dir, "merged_columns.xlsx")
    wb.save(output_path)
    print(f"已保存到: {output_path}")

if __name__ == "__main__":
    main()
