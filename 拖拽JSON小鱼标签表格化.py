#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
json_to_excel.py

将拖拽到此脚本上的 JSON 文件转换为 Excel 表格，所有数据在同一个工作表中，
并自动调整列宽与统一行高。

用法：
    直接将 JSON 文件拖拽到脚本上运行

依赖：
    pip install pandas openpyxl
"""
import os
import sys
import json
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 最大列宽（字符数），超过则换行
MAX_COL_WIDTH = 50
# 统一行高（点数），可根据需要调整
ROW_HEIGHT = 25

def load_json(path):
    try:
        # 使用 utf-8-sig 去除 BOM
        with open(path, "r", encoding="utf-8-sig") as f:
            return json.load(f)
    except Exception as e:
        print(f"加载 JSON 文件时出错: {e}")
        sys.exit(1)


def adjust_column_width_and_wrap(ws, df, max_width=MAX_COL_WIDTH):
    """
    调整列宽并对超出最大宽度的单元格启用自动换行。
    """
    for idx, col in enumerate(df.columns, start=1):
        # 列名长度
        max_length = len(str(col))
        for value in df[col].astype(str):
            if len(value) > max_length:
                max_length = len(value)
        width = min(max_length + 2, max_width)
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = width
        if width >= max_width:
            # 针对该列所有单元格启用换行
            for row in range(1, df.shape[0] + 2):  # 包含表头
                cell = ws[f"{letter}{row}"]
                cell.alignment = Alignment(wrap_text=True)


def set_uniform_row_height(ws, row_count, height=ROW_HEIGHT):
    """
    将所有行设置为统一行高。
    ws: 工作表对象
    row_count: 数据行数 + 1 表头
    height: 行高，单位点
    """
    for row in range(1, row_count + 1):
        ws.row_dimensions[row].height = height


def main(json_path):
    try:
        data = load_json(json_path)

        # 构建记录列表
        records = []
        for url, info in data.get("data", {}).items():
            if url == "meta": continue
            tags = info.get("tags", [])
            meta = info.get("meta", {})
            records.append({
                "url": url,
                "title": meta.get("title", ""),
                "tags": ";".join(tags),
                "created": pd.to_datetime(meta.get("created"), unit="ms", errors="coerce"),
                "updated": pd.to_datetime(meta.get("updated"), unit="ms", errors="coerce"),
            })

        df = pd.DataFrame(records)
        # 输出路径：与 JSON 同目录下，名为 output.xlsx
        out_dir = os.path.dirname(os.path.abspath(json_path))
        out_path = os.path.join(out_dir, "output.xlsx")

        # 写入 Excel
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Data", index=False)
            ws = writer.sheets["Data"]
            # 自动调整列宽并换行
            adjust_column_width_and_wrap(ws, df)
            # 统一行高（包含表头）
            set_uniform_row_height(ws, df.shape[0] + 1)

        print(f"已成功导出到：{out_path}")
    except Exception as e:
        print(f"发生错误：{e}")
        sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("请将 JSON 文件拖拽到此脚本上运行")
        sys.exit(1)

    main(sys.argv[1])
    input("处理完毕，请按任意键退出...")  # 保持窗口打开，方便查看消息
